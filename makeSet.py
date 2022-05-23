import openpyxl


功效集合 = list()
# 發散表寒:0
發散表寒 = {"川烏": 0, "藁本": 0, "蔥": 0, "川芎": 0, "桂枝": 0, "麻黃": 0, "獨活": 0, "防風": 0,
        "紫蘇": 0, "香薷": 0, "蘇梗": 0, "蒼耳子": 0, "生薑": 0, "辛夷": 0, "建神麯": 0, "胡荽": 0, "浮萍": 0}
# 祛風寒:
祛風寒 = {"生薑": 0, "附子": 0, "蒼耳草": 0, "細辛": 0, "桂枝": 0, "川芎": 0}
功效集合.append(發散表寒)
功效集合.append(祛風寒)

# open the excel file
data = openpyxl.load_workbook("原始藥效_執行結果.xlsx")
# deal with the first sheet
sheet = data.worksheets[0]


def checkarr(str):
    if str == '發散表寒':
        return 0
    elif str == '祛風寒':
        return 1
    else:
        return -1

    # ...(可能要寫所有的or有其他方法？)


def checkeffect(int):
    if int == 0:
        return '發散表寒'
    if int == 1:
        return '祛風寒'


# 確認方劑內的大小關係
def check_prescription_order(rowstart, rowfinal):
    effectval = 100
    for row in range(rowstart, rowfinal):
        medicinename = sheet.cell(row, 3).value
        effect = sheet.cell(row, 5).value
        effectindex = checkarr(effect)
        if effectindex != -1:
            # for index in range(len(功效集合[effectIndex]))
            for key in 功效集合[effectindex].keys():
                #medEffect = 功效集合
                if medicinename == key:
                    功效集合[effectindex][key] = effectval
                    effectval -= 10
                    break


# 第一個方劑在第二行(e.g.,from 麻黃湯)
rowStart = 2
for row in range(3, sheet.max_row, 1):
    prescription = sheet.cell(row, 1).value
    # 如果到了下一個方劑(e.g.,桂枝湯)
    if prescription != None:
        rowFinal = row-1
        check_prescription_order(rowStart, rowFinal)
        rowStart = row


# 開新的excel儲存結果
#   A           | B                 |C
# 1 發散表寒    |                    |祛風寒
# 2 川烏        | 100(不知道 我亂打)  |~
# 3 藁本        | 100                |~
# 4 蔥          | 90                 |~


medData = openpyxl.Workbook()
medValSheet = medData.active  # Workbook.create_sheet()
# 功效跟藥材名都放在奇數行
keyIndex = 1
for effectIndex in range(len(功效集合)):
    # 藥效
    medValSheet.cell(1, keyIndex).value = checkeffect(effectIndex)
    # 藥材從第二行開始放
    rowindex = 2
    for key in 功效集合[effectIndex].keys():
        # 藥材名稱
        medValSheet.cell(row=rowindex, column=keyIndex).value = key
        # 效力值(放在偶數行)
        medValSheet.cell(row=rowindex, column=keyIndex +
                         1).value = 功效集合[effectIndex][key]
        rowindex += 1
    keyIndex = keyIndex + 2
    medData.save('效力值.xlsx')

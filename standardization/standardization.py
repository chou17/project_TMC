import openpyxl
from name import 標準功效
from name import 藥材功效

# open excel
data = openpyxl.load_workbook("原始藥效.xlsx")


def systemeffect(row):
    ori_effect = sheet.cell(row, 4).value
    std_col = 5
    bool_system = False
    for check in range(len(標準功效)):
        if ori_effect == 標準功效[check][0]:
            bool_system = True
            for i in range(1, len(標準功效[check])):
                sheet.cell(row, std_col).value = 標準功效[check][i]
                std_col += 1
            break
    return bool_system


def compareeffect(row, index):
    # output is stored at col 5
    std_col = 5

    # original effect's char
    for char_ori in sheet.cell(row, 4).value:
        # std effect
        for str_std in 藥材功效[index]:
            for char_std in str_std:
                # if there is a char is the same between original effect & std effect
                if char_ori == char_std:
                    # if the std effect is repeat at the front cell, continue
                    if(std_col > 5 and str_std == sheet.cell(row, std_col-1).value):
                        continue
                    # else, std effect is stored at col (>=5)
                    sheet.cell(row, std_col).value = str_std
                    std_col += 1


for sheet in data.worksheets:
    # loop the whole file
    for row in range(2, sheet.max_row, 1):
        bool_system = systemeffect(row)

        # medicine's name
        if(bool_system == False):
            checkName = sheet.cell(row, 3).value
            for check in range(len(藥材功效)):
                # 藥材功效[i][0] is the medicine name
                if checkName == 藥材功效[check][0]:
                    compareeffect(row, check)

data.save('原始藥效.xlsx')

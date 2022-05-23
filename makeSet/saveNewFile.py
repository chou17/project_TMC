import openpyxl
from effectSet import 功效集合
from check import checkeffect

# 開新的excel儲存結果
#   A           | B                 |C
# 1 發散表寒    |                    |祛風寒
# 2 川烏        | 100                |~
# 3 藁本        | 100                |~
# 4 蔥          | 90                 |~


medData = openpyxl.Workbook()
medValSheet = medData.active  # Workbook.create_sheet()
# 功效跟藥材名都放在奇數行
keyIndex = 1
for effectIndex in range(len(功效集合)):
    # 藥效
    medValSheet.cell(1, keyIndex).value = checkeffect(effectIndex)
    # 藥材從第二行開始放
    rowindex = 2
    for key in 功效集合[effectIndex].keys():
        # 藥材名稱
        medValSheet.cell(row=rowindex, column=keyIndex).value = key
        # 效力值(放在偶數行)
        medValSheet.cell(row=rowindex, column=keyIndex +
                         1).value = 功效集合[effectIndex][key]
        rowindex += 1
    keyIndex = keyIndex + 2
    medData.save('效力值.xlsx')

from itertools import count
import openpyxl
from collections import defaultdict
from effectSet import 功效集合
from check import checkarr
from check import checkeffect
import time

# 開始測量
start = time.time()

# open the excel file
data = openpyxl.load_workbook("方劑單一藥效.xlsx")


class Graph:
    # construct
    def __init__(self):
        self.graph = defaultdict(list)

    def add_edge(self, u, v):
        # edge<u, v>
        self.graph[u].append(v)

    def g_get_key(self, v):
        for k, val in self.graph.items():
            if val == v:
                return k

    def set_val(self, arr):
        # 有指向其他點的點
        key = list(self.graph.keys())

        up_val = 95
        val = 90

        con = True
        while(con):
            con = False
            val -= 10
            for i in key:
                # 如果有被上一層的人指向，表示為下一層
                if arr[i] == up_val:
                    for j in self.graph[i]:
                        if arr[j] < val:
                            arr[j] = val
                            con = True
            up_val = val
        return val


def check_prescription_order(rowstart, rowfinal, e):
    appear = False
    first_order = ""
    premed = ""
    preorder = ""
    for row in range(rowstart, rowfinal):
        medicinename = sheet.cell(row, 3).value
        effect = sheet.cell(row, 4).value
        effectindex = checkarr(effect)

        if effectindex == e:
            if appear == False:
                appear = True
                first_order = sheet.cell(row, 2).value
                preorder = sheet.cell(row, 2).value
                if preorder == None:
                    print('輸入資料第'+row+'行格式錯誤!')
                    break
                # 方劑中首次出現(方劑中最大功效)
                功效集合[e][medicinename] = 95
                premed = medicinename
            else:
                order = sheet.cell(row, 2).value
                if order == first_order:
                    功效集合[e][medicinename] = 95

                else:
                    if preorder == order:
                        name = g.g_get_key(premed)
                        if name == None:
                            continue
                        premed = name
                    g.add_edge(premed, medicinename)
                    preorder = order

                premed = medicinename


def get_key(dict, val):
    return [k for k, v in dict.items() if v == val]


medData = openpyxl.Workbook()
medValSheet = medData.active  # Workbook.create_sheet()
# 功效跟藥材名都放在奇數行
keyIndex = 1

#str = input("請輸入想要查詢的功效：")
#i = checkarr(str)
# while i <= 221 and i >= 0:
for i in range(222):
    g = Graph()
    for sheet in data.worksheets:
        # 第一個方劑在第二行(e.g.,from 麻黃湯)
        rowStart = 2
        for row in range(3, sheet.max_row, 1):
            prescription = sheet.cell(row, 1).value
            # 如果到了下一個方劑(e.g.,桂枝湯)
            if prescription != None:
                rowFinal = row
                check_prescription_order(rowStart, rowFinal, i)
                rowStart = row
        check_prescription_order(rowStart, sheet.max_row, i)

    val = g.set_val(功效集合[i])
    only_set = get_key(功效集合[i], 0)
    # num = -1代表圖上皆是獨立的點，或者圖中沒有任何點
    # 若圖上皆是獨立的點，only_set必為空(效力值皆為100)，only_set為空則代表圖中沒有任何點

    for only in only_set:
        功效集合[i][only] = val
    if get_key(功效集合[i], 80) != [] and get_key(功效集合[i], 95) == []:
        # print("Excel檔案中無此項功效之資料，無法計算效力值。")
        #str = input("請輸入想要查詢的功效：")
        #i = checkarr(str)
        continue

    # print(功效集合[i])

    # excel 處理
    medValSheet.cell(1, keyIndex).value = checkeffect(i)
    # 藥材從第二行開始放
    rowindex = 2
    for key in 功效集合[i].keys():
        # 藥材名稱
        medValSheet.cell(row=rowindex, column=keyIndex).value = key
        # 效力值(放在偶數行)
        medValSheet.cell(row=rowindex, column=keyIndex +
                         1).value = 功效集合[i][key]
        rowindex += 1
    keyIndex = keyIndex + 2
    # 迴圈
    #str = input("請輸入想要查詢的功效：")
    #i = checkarr(str)

medData.save('效力值.xlsx')
# print("輸入錯誤，程式結束")

end = time.time()

# 輸出結果
print("執行時間：%f 秒" % (end - start))
quit()

from tkinter import *
from tkinter import filedialog
import tkinter as tk
from tkinter import messagebox
import tkinter.font as tkFont
from itertools import count
import openpyxl
from collections import defaultdict
from effectSet import 功效集合
from check import checkarr
from check import checkeffect
import time
import operator
from PIL import ImageTk, Image 


# 開始測量
start = time.time()


class Graph:
    # construct
	def __init__(self):
		self.graph = defaultdict(list)

	def add_edge(self, u, v):
        # edge<u, v>
		self.graph[u].append(v)

	def g_get_key(self, v):
		for k, val in self.graph.items():
			if val == v:
				return k

	def set_val(self, arr):
        # 有指向其他點的點
		key = list(self.graph.keys())
        
		up_val = 95
		val = 90

        # 是否繼續執行
		con = True
		while(con):
			con = False
			val -= 10
			for i in key:
                # 如果有被上一層的人指向，表示為下一層
				if arr[i] == up_val:
					for j in self.graph[i]:
						if arr[j] < val:
							arr[j] = val
							con = True
			up_val -= 10
		return val

class GUI:
	#global variable 
	def __init__():
		self.importFilePath = ""
		self.img
	
	def check_prescription_order(rowstart, rowfinal, e, sheet, g):
		appear = False
		first_order = ""    
		premed = ""
		preorder = ""
		for row in range(rowstart, rowfinal):
			medicinename = sheet.cell(row, 3).value
			effect = sheet.cell(row, 4).value
			effectindex = checkarr(effect)
        
			if effectindex == e:
				if appear == False:
					appear = True
					first_order = sheet.cell(row, 2).value
					preorder = sheet.cell(row, 2).value
					if preorder == None:
						print('輸入資料第'+row+'行格式錯誤!')
						break
                	# 方劑中首次出現(方劑中最大功效)
					功效集合[e][medicinename] = 95
					premed = medicinename
				else:
					order = sheet.cell(row, 2).value
					if order == first_order:
						功效集合[e][medicinename] = 95

					else:
						if preorder == order:
							name = g.g_get_key(premed)
							if name == None:
								continue
							premed = name
						g.add_edge(premed, medicinename)
						preorder = order

					premed = medicinename

	def get_key(dict, val):
		return [k for k, v in dict.items() if v == val]

	#匯入檔案
	def import_file():
		#得到檔案路徑與名稱
		GUI.importFilePath = filedialog.askopenfilename()

	#匯出檔案
	def export_file():
		medData.save(filedialog.askdirectory() + "/效力值.xlsx")
		messagebox.showinfo('匯出檔案', '已匯出\"效力值.xlsx\"至您選擇的資料夾')

	
	#查詢並顯示藥效
	def look_up():
		#獲得選取的藥效
		enames = effectList.curselection() 
		elist = []

		evar = StringVar()
		displayEffect = ""

		for i in enames:	#可多選(目前為單選)
			e = effectList.get(i)
			elist.append(e)
		for evar in elist:	#可多選(目前為單選)
			eindex = checkarr(evar)
			List.delete(0, List.size() + 1)
			List.insert(tk.END, "藥材\t效力值\n")
			for key in 功效集合[eindex].keys():
				List.insert(tk.END, key + "\t" + str(功效集合[eindex][key]) + "\n")

	#更新並計算藥效
	def update_effect():
		# open the excel file
		data = openpyxl.load_workbook(GUI.importFilePath)
		medValSheet = medData.active  # Workbook.create_sheet()
		# 功效跟藥材名都放在奇數行
		keyIndex = 1

		for i in range(0, 224):
			g = Graph()
			for sheet in data.worksheets:
				# 第一個方劑在第二行(e.g.,from 麻黃湯)
				rowStart = 2
				for row in range(3, sheet.max_row, 1):
					prescription = sheet.cell(row, 1).value
         			# 如果到了下一個方劑(e.g.,桂枝湯)
					if prescription != None:
						rowFinal = row
						GUI.check_prescription_order(rowStart, rowFinal, i, sheet, g)
						rowStart = row
				GUI.check_prescription_order(rowStart, sheet.max_row, i, sheet, g)

			val = g.set_val(功效集合[i])
			only_set = GUI.get_key(功效集合[i], 0)
    		# num = -1代表圖上皆是獨立的點，或者圖中沒有任何點
    		# 若圖上皆是獨立的點，only_set必為空(效力值皆為100)，only_set為空則代表圖中沒有任何點
			for only in only_set:
				功效集合[i][only] = val
			if GUI.get_key(功效集合[i], 80) != [] and GUI.get_key(功效集合[i], 95) == []:
				continue

			#排序
			sortedeffect = dict(sorted(功效集合[i].items(), key = operator.itemgetter(1), reverse = True))
			功效集合[i] = sortedeffect

			# excel 處理
			medValSheet.cell(1, keyIndex).value = checkeffect(i)
			# 藥材從第二行開始放
			rowindex = 2
			for key in 功效集合[i].keys():
				# 藥材名稱
				medValSheet.cell(row=rowindex, column=keyIndex).value = key
				# 效力值(放在偶數行)
				medValSheet.cell(row=rowindex, column=keyIndex + 1).value = 功效集合[i][key]
			
				rowindex += 1
			keyIndex = keyIndex + 2
			#加入清單
			effectList.insert(tk.END, checkeffect(i))
	

#開新的excel以儲存效力值
medData = openpyxl.Workbook()

#視窗設定
window = tk.Tk()
window.title('基於中藥方劑配伍結構的藥材療效效力值推論系統')
screen_width = window.winfo_screenwidth()
screen_height = window.winfo_screenheight()
window.geometry(str(screen_width) + 'x' + str(screen_height))
window.resizable(True,True)
#插入背景圖片
bg = Image.open('專題.jpg')
photo = ImageTk.PhotoImage(bg)
canvas = tk.Canvas(window, width = bg.size[0], height = bg.size[1])
canvas.pack()
canvas.create_image(0, 0, anchor = tk.NW, image = photo)

#簡介
label = Message(window, text = "藥材效力值計算", justify = LEFT, fg = '#337781', bg = '#fff', font = tkFont.Font(family = "Microsoft Yahei", size = 36), width = 600, padx = 10, pady = 10)
label.place(relx = 0.025, rely = 0.35)

#使用說明
label = Message(window, text = "小叮嚀：使用本系統前，請先將方劑資料儲存為excel檔，格式如右圖    👉   👉    \n\n使用說明：\n步驟一 ▼\n請點選\"匯入檔案\"，開啟欲匯入之excel檔\n步驟二 ▼\n點選\"更新\"按鈕\n步驟三 ▼\n◆查詢單一功效之藥材效力值：從下方選單中選取欲查詢之功效\n◆匯出所有功效之藥材效力值(excel檔):點選\"匯出檔案\"，選擇欲儲存檔案的資料夾", 
	justify = LEFT, bd = 10, bg = '#337781', fg = '#fff', font = tkFont.Font(family = "Microsoft Yahei", size = 14), width = 700)
label.place(relx = 0.275, rely = 0.025)

#"匯入"按鈕設定
iButton = tk.Button(text = "匯入檔案", command = GUI.import_file, font = tkFont.Font(family = "Microsoft Yahei", size = 14))
iButton.place(relx = 0.525, rely = 0.4, anchor = CENTER)

#藥效選單設定
effectName = tk.StringVar()
effectList = tk.Listbox(window, listvariable = effectName, selectmode = SINGLE, font = tkFont.Font(family = "Microsoft Yahei", size = 14))

effectList.yview()
effectList.yview_scroll(1,UNITS)
effectList.place(relx = 0.575, rely = 0.6, anchor = CENTER)

#藥效選單更新按鈕設定
lButton = tk.Button(text = "更新", command = GUI.update_effect, font = tkFont.Font(family = "Microsoft Yahei", size = 14))
lButton.place(relx = 0.65, rely = 0.4, anchor = CENTER)

#單一藥效顯示畫面設定
effectvar = StringVar()
List = tk.Listbox(window, listvariable = effectvar, exportselection = 0, height = 16, font = tkFont.Font(family = "Microsoft Yahei", size = 14))
List.insert(tk.END, "---請選擇欲查詢藥效後---")
List.insert(tk.END, "---點擊下方\"查詢\"按鈕---")

List.yview()
List.yview_scroll(1,UNITS)
List.place(relx = 0.75, rely = 0.375)

#單一藥效選擇按鈕設定
eButton = tk.Button(text = "查詢", command = GUI.look_up, font = tkFont.Font(family = "Microsoft Yahei", size = 14))
eButton.place(relx = 0.65, rely = 0.8, anchor = CENTER)

#"匯出"按鈕設定
iButton = tk.Button(text = "匯出檔案", command = GUI.export_file, font = tkFont.Font(family = "Microsoft Yahei", size = 14))
iButton.place(relx = 0.525, rely = 0.8, anchor = CENTER)

#create
label = Message(window, text = "created by Yuhsuan, Jessie", bg = '#75C8C8', justify = LEFT, font = tkFont.Font(family = "Microsoft Yahei", size = 14), width = 600)
label.place(relx = 0.4125, rely = 0.95)

#維持程式(視窗)運作
window.mainloop()

#計時結束
end = time.time()
# 輸出結果
print("執行時間：%f 秒" % (end - start))
quit()
from tkinter import *
from tkinter import filedialog
import tkinter as tk
from tkinter import messagebox
import tkinter.font as tkFont
from itertools import count
import openpyxl
from collections import defaultdict
from effectSet import 功效集合
from check import checkarr
from check import checkeffect
import time
import operator
from PIL import ImageTk, Image 


# 開始測量
start = time.time()


class Graph:
    # construct
	def __init__(self):
		self.graph = defaultdict(list)

	def add_edge(self, u, v):
        # edge<u, v>
		self.graph[u].append(v)
        #取得key
	def g_get_key(self, v):
		for k, val in self.graph.items():
			if val == v:
				return k

	def set_val(self, arr):
        # key:有指向其他點的點
		key = list(self.graph.keys())
        #效力值最大值=95
		up_val = 95
		val = 90

        # 是否繼續執行
		con = True
		while(con):
			con = False
			#效力值遞減(-10)
			val -= 10
			for i in key:
                # 若值=up_val，表有被指向，即為下一層
			    if arr[i] == up_val:
				    #所有下一層的元素
				    for j in self.graph[i]:
					    #若沒有被寫入值過
                        #有被寫過則代表出現a>=b, b>=a之情況，保留較大的值以表a=b
				        if arr[j] < val:
				            arr[j] = val
                            con = True
            up_val -= 10
        return val

class GUI:
	#global variable 
	def __init__():
		self.importFilePath = ""
		self.img
	
	#根據方劑配伍結構畫出對應的有向圖
	def check_prescription_order(rowstart, rowfinal, e, sheet, g):
		appear = False
		#此功效之最大配伍結構
        first_order = ""    
        #parent藥材名稱
		premed = ""
        #parent方劑配伍結構
		preorder = ""
        #分析單一方劑
		for row in range(rowstart, rowfinal):
			medicinename = sheet.cell(row, 3).value	#取得藥材名稱
			effect = sheet.cell(row, 4).value	#取得療效名稱
			effectindex = checkarr(effect)	#將療效名稱轉換成相對應的index
        	
        	#如果藥材療效跟欲比對之療效相同
			if effectindex == e:
				if appear == False:	#有此療效之藥材在此方劑中首次出現
					appear = True
					#記錄第一個配伍結構(head)
					first_order = sheet.cell(row, 2).value
					preorder = sheet.cell(row, 2).value
					if preorder == None:
						print('輸入資料第'+row+'行格式錯誤!')
						break
                	# 方劑中首次出現(方劑中最大功效)
					功效集合[e][medicinename] = 95
                    #給完值後紀錄為parent
					premed = medicinename
				
				else:	#在此方劑已出現過有此療效之藥材
					order = sheet.cell(row, 2).value
					if order == first_order:	#如果當前藥材與head配伍結構相同
						功效集合[e][medicinename] = 95	#效力值設為最高

					#如果當前藥材跟head配伍結構不同(即為children)
					else:	
						#如果當前藥材跟parent配伍結構相同
						if preorder == order:
							name = g.g_get_key(premed)	#取得grandparent藥材名稱
							if name == None:	#如果grandparent不存在
								continue	#繼續執行
							
							#如果grandparent存在
							premed = name	#將grandparent設為premed，由grandparent指向當前藥材
						#添加edge
						g.add_edge(premed, medicinename)	#效力較高者指向效力較低者
					    #更新
						preorder = order	

					premed = medicinename

	def get_key(dict, val):
		return [k for k, v in dict.items() if v == val]

	#匯入檔案
	def import_file():
		#得到檔案路徑與名稱
		GUI.importFilePath = filedialog.askopenfilename()

	#匯出檔案
	def export_file():
		medData.save(filedialog.askdirectory() + "/效力值.xlsx")
		messagebox.showinfo('匯出檔案', '已匯出\"效力值.xlsx\"至您選擇的資料夾')

	
	#查詢並顯示藥效
	def look_up():
		#獲得選取的藥效
		enames = effectList.curselection() 
		elist = []

		evar = StringVar()
		displayEffect = ""

		#取得使用者選擇的療效
		for i in enames:	#可多選(目前為單選)
			e = effectList.get(i)	
			elist.append(e)

		#顯示具使用者選擇的療效之藥材及效力值
		for evar in elist:	#可多選(目前為單選)
			eindex = checkarr(evar)
			List.delete(0, List.size() + 1)	#清空前一次顯示
			List.insert(tk.END, "藥材：效力值\n")
			for key in 功效集合[eindex].keys():
				List.insert(tk.END, key + "：" + str(功效集合[eindex][key]) + "\n")

	#更新並計算藥效
	def update_effect():
		# open the excel file
		data = openpyxl.load_workbook(GUI.importFilePath)
		medValSheet = medData.active  # Workbook.create_sheet()
		# 功效跟藥材名都放在奇數行
		keyIndex = 1

		for i in range(0, 224):
			g = Graph()
			for sheet in data.worksheets:
				# 第一個方劑在第二行(e.g.,from 麻黃湯)
				rowStart = 2
				for row in range(3, sheet.max_row, 1):
					prescription = sheet.cell(row, 1).value
         			# 如果到了下一個方劑(e.g.,桂枝湯)
					if prescription != None:
						rowFinal = row
						GUI.check_prescription_order(rowStart, rowFinal, i, sheet, g)	#根據該方劑取得偏序關係
						rowStart = row
				GUI.check_prescription_order(rowStart, sheet.max_row, i, sheet, g)	#整理全部方劑所得的偏序關係(整理出整個有向圖)
            #取得最小效力值
			val = g.set_val(功效集合[i])
			#尚未被賦值的點
            only_set = GUI.get_key(功效集合[i], 0)
    		#賦最小效力值
			for only in only_set:
				功效集合[i][only] = val
            #若該療效有效力值80卻沒有效力值95，表示資料中沒有此療效
			if GUI.get_key(功效集合[i], 80) != [] and GUI.get_key(功效集合[i], 95) == []:
				#繼續執行
                continue

			#排序
			sortedeffect = dict(sorted(功效集合[i].items(), key = operator.itemgetter(1), reverse = True))
			功效集合[i] = sortedeffect

			# excel 處理
			medValSheet.cell(1, keyIndex).value = checkeffect(i)
			# 藥材從第二行開始放
			rowindex = 2
			for key in 功效集合[i].keys():
				# 藥材名稱
				medValSheet.cell(row=rowindex, column=keyIndex).value = key
				# 效力值(放在偶數行)
				medValSheet.cell(row=rowindex, column=keyIndex + 1).value = 功效集合[i][key]
			
				rowindex += 1
			keyIndex = keyIndex + 2
			#加入清單
			effectList.insert(tk.END, checkeffect(i))
	

#開新的excel以儲存效力值
medData = openpyxl.Workbook()

#視窗設定
window = tk.Tk()
window.title('基於中藥方劑配伍結構的藥材療效效力值推論系統')
screen_width = window.winfo_screenwidth()
screen_height = window.winfo_screenheight()
window.geometry(str(screen_width) + 'x' + str(screen_height))
window.resizable(True,True)
#插入背景圖片
bg = Image.open('專題.jpg')
photo = ImageTk.PhotoImage(bg)
canvas = tk.Canvas(window, width = bg.size[0], height = bg.size[1])
canvas.pack()
canvas.create_image(0, 0, anchor = tk.NW, image = photo)

#簡介
label = Message(window, text = "藥材效力值計算", justify = LEFT, fg = '#337781', bg = '#fff', font = tkFont.Font(family = "Microsoft Yahei", size = 36), width = 600, padx = 10, pady = 10)
label.place(relx = 0.025, rely = 0.35)

#使用說明
label = Message(window, text = "小叮嚀：使用本系統前，請先將方劑資料儲存為excel檔，格式如右圖    👉   👉    \n\n使用說明：\n步驟一 ▼\n請點選\"匯入檔案\"，開啟欲匯入之excel檔\n步驟二 ▼\n點選\"更新\"按鈕\n步驟三 ▼\n◆查詢單一功效之藥材效力值：從下方選單中選取欲查詢之功效\n◆匯出所有功效之藥材效力值(excel檔):點選\"匯出檔案\"，選擇欲儲存檔案的資料夾", 
	justify = LEFT, bd = 10, bg = '#337781', fg = '#fff', font = tkFont.Font(family = "Microsoft Yahei", size = 14), width = 700)
label.place(relx = 0.275, rely = 0.025)

#"匯入"按鈕設定
iButton = tk.Button(text = "匯入檔案", command = GUI.import_file, font = tkFont.Font(family = "Microsoft Yahei", size = 14))
iButton.place(relx = 0.525, rely = 0.4, anchor = CENTER)

#藥效選單設定
effectName = tk.StringVar()
effectList = tk.Listbox(window, listvariable = effectName, selectmode = SINGLE, font = tkFont.Font(family = "Microsoft Yahei", size = 14))

effectList.yview()
effectList.yview_scroll(1,UNITS)
effectList.place(relx = 0.575, rely = 0.6, anchor = CENTER)

#藥效選單更新按鈕設定
lButton = tk.Button(text = "更新", command = GUI.update_effect, font = tkFont.Font(family = "Microsoft Yahei", size = 14))
lButton.place(relx = 0.65, rely = 0.4, anchor = CENTER)

#單一藥效顯示畫面設定
effectvar = StringVar()
List = tk.Listbox(window, listvariable = effectvar, exportselection = 0, height = 16, font = tkFont.Font(family = "Microsoft Yahei", size = 14))
List.insert(tk.END, "---請選擇欲查詢藥效後---")
List.insert(tk.END, "---點擊下方\"查詢\"按鈕---")

List.yview()
List.yview_scroll(1,UNITS)
List.place(relx = 0.75, rely = 0.375)

#單一藥效選擇按鈕設定
eButton = tk.Button(text = "查詢", command = GUI.look_up, font = tkFont.Font(family = "Microsoft Yahei", size = 14))
eButton.place(relx = 0.65, rely = 0.8, anchor = CENTER)

#"匯出"按鈕設定
iButton = tk.Button(text = "匯出檔案", command = GUI.export_file, font = tkFont.Font(family = "Microsoft Yahei", size = 14))
iButton.place(relx = 0.525, rely = 0.8, anchor = CENTER)

#create
label = Message(window, text = "created by Yuhsuan, Tzuhsuan", bg = '#75C8C8', justify = LEFT, font = tkFont.Font(family = "Microsoft Yahei", size = 14), width = 600)
label.place(relx = 0.4125, rely = 0.95)

#維持程式(視窗)運作
window.mainloop()

#計時結束
end = time.time()
# 輸出結果
print("執行時間：%f 秒" % (end - start))
quit()

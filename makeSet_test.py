import openpyxl

功效集合 = [
    # 發散表寒:0
    {"川烏": 0, "藁本": 0, "蔥": 0, "川芎": 0, "桂枝": 0, "麻黃": 0, "獨活": 0, "防風": 0, "紫蘇": 0,
        "香薷": 0, "蘇梗": 0, "蒼耳子": 0, "生薑": 0, "辛夷": 0, "建神麯": 0, "胡荽": 0, "浮萍": 0},
 
]
#open the excel file
data = openpyxl.load_workbook("原始藥效_執行結果.xlsx")
#deal with the first sheet
sheet = data.worksheets[0]

def checkarr(str):
    if str == '發散表寒':
        return 0
    elif str == '祛風寒':
        return 1
    #...(可能要寫所有的or有其他方法？)

def checkeffect(int):
    if int == 0 :
        return '發散表寒'


#確認方劑內的大小關係    
def check_prescription_order(rowStart,rowFinal):
    effectVal = 100
    for row in range(rowStart, rowFinal):
        medicineName = sheet.cell(row, 3).value
        effect = sheet.cell(row, 5).value
        effectIndex = checkarr(effect)
        for key in 功效集合[effectIndex].keys():
            if medicineName == key:
                功效集合[effectIndex][key] = effectVal
                effectVal -= 10
                break
    
#第一個方劑在第二行(e.g.,from 麻黃湯)
rowStart = 2
for row in range(3, sheet.max_row, 1):
    prescription = sheet.cell(row, 1).value
    #如果到了下一個方劑(e.g.,桂枝湯)
    if(prescription != None):
        rowFinal = row-1
        check_prescription_order(rowStart,rowFinal)
        rowStart = row
        
        
#開新的excel儲存結果
#   A           | B                 |C
# 1 發散表寒    |                    |祛風寒
# 2 川烏        | 100(不知道 我亂打)  |~
# 3 藁本        | 100                |~
# 4 蔥          | 90                 |~


medData = openpyxl.Workbook()
medValSheet = medData.active   #Workbook.create_sheet()
for effectIndex in range(len(功效集合)):
    #print(effectIndex)
    keyIndex = 1
    medValSheet.cell(row = effectIndex,column = keyIndex).value = checkeffect(effectIndex)
    keyIndex = keyIndex + 1
    for key in 功效集合[effectIndex].keys():
        medValSheet.cell(row = effectIndex+1, column = keyIndex).value = key
        medValSheet.cell(row = effectIndex+1, column = keyIndex + 1).value = 功效集合[effectIndex][key]
        keyIndex = keyIndex + 1
    medData.save('效力值.xlsx')
